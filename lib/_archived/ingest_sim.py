import logging
import re
from datetime import datetime

import pandas as pd

from lib.db import init_db, get_connection
from lib.file_manager import scan_sims, rename_sim
from lib.normalizers import normalize_wallet, normalize_token_id, normalize_game

logger = logging.getLogger(__name__)


def _find_header_row(ws, target_cell_value):
    """Find the row index containing target_cell_value in the first column."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True)):
        if row[0] and str(row[0]).strip() == target_cell_value:
            return i
    return None


def _parse_results_sheet(wb):
    """Extract wallet data from 📊 Results sheet."""
    ws = wb['📊 Results']

    # Find header row
    header_idx = _find_header_row(ws, 'Wallet Address')
    if header_idx is None:
        logger.error("Could not find 'Wallet Address' header in Results sheet")
        return []

    rows = list(ws.iter_rows(min_row=header_idx + 1, values_only=True))
    header = rows[0]
    col_map = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    wallets = []
    for row in rows[1:]:
        addr = row[col_map.get('Wallet Address', 0)]
        if not addr or str(addr).strip() == '':
            continue

        def g(name, default=None):
            idx = col_map.get(name)
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        wallets.append({
            'wallet_address': normalize_wallet(str(addr)),
            'category': g('Category'),
            'subcategory': g('Subcategory'),
            'detail': normalize_game(str(g('Detail', '')), source='sim_detail'),
            'trades': g('Trades'),
            'sim_trades': g('Sim Trades'),
            'volume': g('Volume'),
            'sim_pnl': g('💰 Sim PnL'),
            'sim_roi_pct': g('📈 Sim ROI %'),
            'max_drawdown_pct': g('📉 Max DD %'),
            'copied': g('✅ Copied'),
            'skipped': g('⏭️ Skipped'),
            'peak_outflow_30d': g('📤 Peak Out 30d'),
            'lb_all_time': g('LB All-Time $'),
            'lb_name': g('LB Name'),
            'gamma_cash_pnl': g('Gamma Cash PnL'),
        })

    return wallets


def _parse_sum_sheet(wb, sheet_name):
    """Parse a _SUM sheet for P&L concentration data."""
    try:
        ws = wb[sheet_name]
    except KeyError:
        return None

    # Find header row (search for "Market ID")
    header_idx = _find_header_row(ws, 'Market ID')
    if header_idx is None:
        return None

    rows = list(ws.iter_rows(min_row=header_idx + 1, values_only=True))
    header = rows[0]
    col_map = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    pnl_col = col_map.get('Total PnL $')
    if pnl_col is None:
        return None

    market_pnls = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        pnl_val = row[pnl_col] if pnl_col < len(row) else None
        if pnl_val is not None:
            try:
                market_pnls.append(float(pnl_val))
            except (ValueError, TypeError):
                continue

    if not market_pnls:
        return None

    # Sort by absolute value descending
    sorted_pnls = sorted(market_pnls, key=lambda x: abs(x), reverse=True)
    total_abs = sum(abs(p) for p in sorted_pnls)
    total_positive = sum(p for p in market_pnls if p > 0)

    if total_abs == 0:
        return {'top1': 0, 'top3': 0, 'ohw': 0}

    top1 = abs(sorted_pnls[0]) / total_abs if sorted_pnls else 0
    top3 = sum(abs(p) for p in sorted_pnls[:3]) / total_abs if len(sorted_pnls) >= 3 else top1
    ohw = max(market_pnls) / total_positive if total_positive > 0 else 0

    return {'top1': top1, 'top3': top3, 'ohw': ohw}


def _parse_drl_sheet(wb, sheet_name):
    """Parse a _DRL sheet for behavioral profile data."""
    try:
        ws = wb[sheet_name]
    except KeyError:
        return None

    # Find header row dynamically
    header_idx = _find_header_row(ws, 'Timestamp (UTC)')
    if header_idx is None:
        return None

    # Read all rows from header onward
    all_rows = list(ws.iter_rows(min_row=header_idx + 1, values_only=True))
    if not all_rows:
        return None

    header = all_rows[0]
    col_map = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    # Required columns
    price_col = col_map.get('Source Price')
    token_col = col_map.get('Token ID')
    market_col = col_map.get('Market ID')
    side_col = col_map.get('Side')
    shares_col = col_map.get('Source Shares')
    if price_col is None:
        return None

    prices = []
    # market_id -> {token_id: total_buy_shares}
    market_buy_shares = {}

    for row in all_rows[1:]:
        if row[0] is None:
            continue

        # Source Price
        if price_col < len(row) and row[price_col] is not None:
            try:
                prices.append(float(row[price_col]))
            except (ValueError, TypeError):
                pass

        # Track buy shares per token per market for arb detection
        if market_col is not None and token_col is not None and side_col is not None and shares_col is not None:
            mid = row[market_col] if market_col < len(row) else None
            tid = row[token_col] if token_col < len(row) else None
            side = row[side_col] if side_col < len(row) else None
            shares = row[shares_col] if shares_col < len(row) else None

            if mid and tid and side and str(side).strip().upper() == 'BUY' and shares:
                mid_str = str(mid).strip()
                tid_str = str(tid).strip()
                try:
                    share_val = float(shares)
                except (ValueError, TypeError):
                    continue
                if mid_str not in market_buy_shares:
                    market_buy_shares[mid_str] = {}
                if tid_str not in market_buy_shares[mid_str]:
                    market_buy_shares[mid_str][tid_str] = 0.0
                market_buy_shares[mid_str][tid_str] += share_val

    if not prices:
        return None

    total_trades = len(prices)
    median_price = sorted(prices)[len(prices) // 2]
    mean_price = sum(prices) / len(prices)
    above_95 = sum(1 for p in prices if p >= 0.95)
    pct_above_95 = above_95 / total_trades

    # Arb detection: buying both sides of the same market where smaller side >= 50% of larger side
    arb_markets = 0
    total_markets = len(market_buy_shares)
    for mid, token_shares in market_buy_shares.items():
        if len(token_shares) >= 2:
            # Has buys on 2+ outcomes — check size ratio
            sorted_shares = sorted(token_shares.values(), reverse=True)
            larger = sorted_shares[0]
            smaller = sorted_shares[1]
            if larger > 0 and smaller >= larger * 0.5:
                arb_markets += 1
    both_sides_pct = arb_markets / total_markets if total_markets > 0 else 0

    return {
        'median_entry_price': median_price,
        'mean_entry_price': mean_price,
        'pct_entries_above_95': pct_above_95,
        'unique_markets': total_markets,
        'total_trades': total_trades,
        'market_diversity_ratio': total_markets / total_trades if total_trades > 0 else 0,
        'both_sides_market_pct': both_sides_pct,
        'has_arb_pattern': 1 if both_sides_pct > 0.1 else 0,
        'has_scalp_pattern': 1 if pct_above_95 > 0.3 else 0,
    }


def run():
    """Ingest all unprocessed sharp sim xlsx files."""
    init_db()
    conn = get_connection()

    files = scan_sims()
    if not files:
        print("No unprocessed sharp sim files found in data/sims/")
        conn.close()
        return

    for filepath in files:
        print(f"Processing {filepath.name}...")

        # Get next sim number
        sim_num = conn.execute(
            "SELECT COALESCE(MAX(sim_number), 0) + 1 FROM sim_registry"
        ).fetchone()[0]

        # Rename file
        new_path = rename_sim(filepath, sim_num)
        sim_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Load workbook
        import openpyxl
        try:
            wb = openpyxl.load_workbook(new_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ERROR loading {new_path.name}: {e}")
            continue

        # Parse Results sheet
        wallets = _parse_results_sheet(wb)
        if not wallets:
            print(f"  No wallets found in Results sheet")
            continue

        # Register sim
        conn.execute("""
            INSERT INTO sim_registry (original_filename, renamed_filename, sim_date, wallet_count)
            VALUES (?, ?, ?, ?)
        """, (filepath.name, new_path.name, sim_date, len(wallets)))

        # Insert snapshots
        for w in wallets:
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO sim_snapshots
                    (sim_number, wallet_address, category, subcategory, detail,
                     trades, sim_trades, volume, sim_pnl, sim_roi_pct,
                     max_drawdown_pct, copied, skipped, peak_outflow_30d,
                     lb_all_time, lb_name, gamma_cash_pnl)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    sim_num, w['wallet_address'], w['category'], w['subcategory'],
                    w['detail'], w['trades'], w['sim_trades'], w['volume'],
                    w['sim_pnl'], w['sim_roi_pct'], w['max_drawdown_pct'],
                    w['copied'], w['skipped'], w['peak_outflow_30d'],
                    w['lb_all_time'], w['lb_name'], w['gamma_cash_pnl'],
                ))
            except Exception as e:
                logger.warning("Failed to insert sim snapshot for %s: %s", w['wallet_address'], e)

        # Parse SUM/DRL sheets for profiles
        complete = 0
        incomplete = 0
        sheet_names = wb.sheetnames

        # Find wallet sheet patterns: NN_0xABCDEF_GAME_SUM / _DRL
        wallet_sheets = {}
        for sn in sheet_names:
            match = re.match(r'(\d+)_(0x[a-fA-F0-9]+)_(\w+)_(SUM|DRL)', sn)
            if match:
                key = (match.group(1), match.group(2).lower())
                if key not in wallet_sheets:
                    wallet_sheets[key] = {}
                wallet_sheets[key][match.group(4)] = sn

        for (num, wallet_prefix), sheets in wallet_sheets.items():
            wallet_addr = None
            for w in wallets:
                if w['wallet_address'].startswith(wallet_prefix):
                    wallet_addr = w['wallet_address']
                    break
            if not wallet_addr:
                wallet_addr = wallet_prefix

            sum_data = None
            drl_data = None
            profile_complete = 1

            if 'SUM' in sheets:
                try:
                    sum_data = _parse_sum_sheet(wb, sheets['SUM'])
                except Exception as e:
                    logger.warning("Failed to parse SUM sheet %s: %s", sheets['SUM'], e)

            if 'DRL' in sheets:
                try:
                    drl_data = _parse_drl_sheet(wb, sheets['DRL'])
                except Exception as e:
                    logger.warning("Failed to parse DRL sheet %s: %s", sheets['DRL'], e)

            if sum_data is None or drl_data is None:
                profile_complete = 0
                incomplete += 1
            else:
                complete += 1

            # Find game from wallet detail
            detail = None
            for w in wallets:
                if w['wallet_address'] == wallet_addr:
                    detail = w['detail']
                    break

            try:
                conn.execute("""
                    INSERT OR REPLACE INTO sim_profiles
                    (sim_number, wallet_address, detail, profile_complete,
                     median_entry_price, mean_entry_price, pct_entries_above_95,
                     pnl_concentration_top1, pnl_concentration_top3,
                     unique_markets, total_trades, market_diversity_ratio,
                     both_sides_market_pct, one_hit_wonder_score,
                     has_arb_pattern, has_scalp_pattern)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    sim_num, wallet_addr, detail, profile_complete,
                    drl_data['median_entry_price'] if drl_data else None,
                    drl_data['mean_entry_price'] if drl_data else None,
                    drl_data['pct_entries_above_95'] if drl_data else None,
                    sum_data['top1'] if sum_data else None,
                    sum_data['top3'] if sum_data else None,
                    drl_data['unique_markets'] if drl_data else None,
                    drl_data['total_trades'] if drl_data else None,
                    drl_data['market_diversity_ratio'] if drl_data else None,
                    drl_data['both_sides_market_pct'] if drl_data else None,
                    sum_data['ohw'] if sum_data else None,
                    drl_data['has_arb_pattern'] if drl_data else 0,
                    drl_data['has_scalp_pattern'] if drl_data else 0,
                ))
            except Exception as e:
                logger.warning("Failed to insert profile for %s: %s", wallet_addr, e)

        conn.commit()
        wb.close()

        print(f"  Ingested Sharp Sim #{sim_num} ({filepath.name}) with {len(wallets)} wallets "
              f"({complete} complete profiles, {incomplete} incomplete)")

    conn.close()
