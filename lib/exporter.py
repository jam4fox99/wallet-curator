"""Wallet data export: XLSX report and TXT wallet list."""
import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from lib.time_utils import now_utc, parse_db_timestamp

logger = logging.getLogger(__name__)

GREEN_FONT = Font(color="22c55e", bold=True)
RED_FONT = Font(color="ef4444", bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")


def _pnl_font(value):
    if value is None:
        return Font()
    return GREEN_FONT if value > 0 else RED_FONT if value < 0 else Font()


def _write_header(ws, columns):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


def export_xlsx(conn) -> bytes:
    """Build XLSX workbook with Wallet Summary and Promotion History sheets."""
    wb = Workbook()

    # ── Sheet 1: Wallet Summary ──────────────────────────────────
    ws = wb.active
    ws.title = "Wallet Summary"

    columns = [
        "Wallet", "Game", "In CSV", "Tier", "Copy %",
        "Realized", "Unrealized", "Total P&L", "ROI %", "Invested",
        "Markets", "Trades", "Days Active",
        "First Trade", "Last Trade",
        "Tier Assigned", "Days In Tier",
        "At-Promo P&L", "Since-Promo P&L",
        "Excluded Positions",
    ]
    _write_header(ws, columns)

    # Build tier lookup
    tier_rows = conn.execute(
        "SELECT wallet_address, tier_name, assigned_at FROM wallet_tiers"
    ).fetchall()
    tier_map = {r["wallet_address"]: r for r in tier_rows}

    # Build tier config lookup
    tier_configs = conn.execute("SELECT * FROM tier_config").fetchall()
    tier_config_map = {r["tier_name"]: float(r["copy_percentage"] or 0) for r in tier_configs}

    # Build game filter lookup
    game_rows = conn.execute(
        "SELECT wallet_address, game_filter FROM synced_active_wallets"
    ).fetchall()
    game_map = {r["wallet_address"]: r["game_filter"] for r in game_rows}

    # Build at-promo P&L lookup (latest promotion per wallet)
    promo_rows = conn.execute(
        """
        SELECT DISTINCT ON (wallet_address) wallet_address, total_pnl_at_action
        FROM promotion_history
        WHERE action IN ('added', 'promoted', 'demoted')
        ORDER BY wallet_address, action_at DESC, id DESC
        """
    ).fetchall()
    at_promo_map = {r["wallet_address"]: float(r["total_pnl_at_action"] or 0) for r in promo_rows}

    # Get all wallets with P&L
    wallets = conn.execute(
        "SELECT * FROM wallet_pnl ORDER BY total_pnl DESC"
    ).fetchall()

    now = now_utc()
    row_idx = 2
    for w in wallets:
        addr = w["master_wallet"]
        tier_info = tier_map.get(addr)
        tier_name = tier_info["tier_name"] if tier_info else ""
        copy_pct = tier_config_map.get(tier_name, 0) if tier_name else None

        # Days active
        first_trade = None
        days_active = 0
        if w["first_trade"]:
            try:
                first_trade = parse_db_timestamp(w["first_trade"])
                days_active = max((now - first_trade).days, 0)
            except Exception:
                pass

        # Tier assignment
        assigned_at = None
        days_in_tier = 0
        if tier_info and tier_info["assigned_at"]:
            try:
                assigned_at = parse_db_timestamp(str(tier_info["assigned_at"]))
                days_in_tier = max((now - assigned_at).days, 0)
            except Exception:
                pass

        # ROI
        invested = float(w["total_invested"] or 0)
        total_pnl = float(w["total_pnl"] or 0)
        roi = round((total_pnl / invested) * 100, 2) if invested > 0 else 0

        # Since-promo P&L
        at_promo_pnl = at_promo_map.get(addr)
        since_promo = round(total_pnl - at_promo_pnl, 2) if at_promo_pnl is not None else None

        game = game_map.get(addr) or w["game"] or ""

        in_csv = "Yes" if addr in game_map else "No"

        values = [
            addr,
            game,
            in_csv,
            tier_name.replace("_", " ").title() if tier_name else "",
            copy_pct,
            round(float(w["realized_pnl"] or 0), 2),
            round(float(w["unrealized_pnl"] or 0), 2),
            round(total_pnl, 2),
            roi,
            round(invested, 2),
            int(w["unique_markets"] or 0),
            int(w["total_trades"] or 0),
            days_active,
            str(w["first_trade"] or "")[:19],
            str(w["last_trade"] or "")[:19],
            str(tier_info["assigned_at"] or "")[:19] if tier_info else "",
            days_in_tier,
            round(at_promo_pnl, 2) if at_promo_pnl is not None else None,
            since_promo,
            int(w["excluded_positions"] or 0),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Color P&L columns (shifted +1 for In CSV column)
            if col_idx in (6, 7, 8, 18, 19) and isinstance(val, (int, float)):
                cell.font = _pnl_font(val)
                cell.number_format = '#,##0.00'
            elif col_idx in (5, 9):
                cell.number_format = '0.00'
            elif col_idx == 10:
                cell.number_format = '#,##0.00'
        row_idx += 1

    _auto_width(ws)

    # ── Sheet 2: Promotion History ───────────────────────────────
    ws2 = wb.create_sheet("Promotion History")

    promo_columns = [
        "Wallet", "Action", "From Tier", "To Tier", "Date",
        "P&L At Action", "Invested At Action",
        "Markets", "Trades", "Days Active", "ROI %",
        "Old Copy %", "New Copy %",
    ]
    _write_header(ws2, promo_columns)

    all_promos = conn.execute(
        """
        SELECT * FROM promotion_history
        ORDER BY action_at DESC, id DESC
        """
    ).fetchall()

    row_idx = 2
    for p in all_promos:
        values = [
            p["wallet_address"],
            str(p["action"] or "").title(),
            str(p["from_tier"] or "").replace("_", " ").title(),
            str(p["to_tier"] or "").replace("_", " ").title(),
            str(p["action_at"] or "")[:19],
            round(float(p["total_pnl_at_action"] or 0), 2),
            round(float(p["total_invested_at_action"] or 0), 2),
            int(p["unique_markets_at_action"] or 0),
            int(p["total_trades_at_action"] or 0),
            int(p["days_active_at_action"] or 0),
            round(float(p["roi_pct_at_action"] or 0), 2),
            p["old_copy_pct"],
            p["new_copy_pct"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 6 and isinstance(val, (int, float)):
                cell.font = _pnl_font(val)
                cell.number_format = '#,##0.00'
        row_idx += 1

    _auto_width(ws2)

    # ── Save to bytes ────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_wallet_list_txt(conn) -> str:
    """Return plain text with one wallet address per line for every traded wallet."""
    rows = conn.execute(
        "SELECT DISTINCT master_wallet FROM trades ORDER BY master_wallet"
    ).fetchall()
    return "\n".join(r["master_wallet"] for r in rows) + "\n"
