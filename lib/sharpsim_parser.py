from __future__ import annotations

from collections import Counter
from datetime import date
from io import BytesIO
from datetime import timedelta
from typing import Any

from openpyxl import load_workbook

from lib.clickhouse_charts import (
    CURATION_ALL_RANGE,
    CURATION_RANGE_DAYS,
    build_chart_payload,
    normalize_curation_range_key,
)


_RESULTS_REQUIRED = {
    "Wallet Address",
    "Category",
    "Subcategory",
    "Detail",
    "Trades",
    "Volume",
    "💰 Sim 1d",
    "💰 Sim 7d",
    "💰 Sim 30d",
    "📈 Sim ROI %",
    "📉 Max DD %",
    "✅ Copied",
    "⏭️ Skipped",
}

_DRL_REQUIRED = {
    "Timestamp (UTC)",
    "Status",
    "Side",
    "Market ID",
    "Question",
    "Token ID",
    "Copied Price",
    "Copied Shares",
    "Copied Notional $",
    "Copied Fee $",
}


def _header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value).strip(): idx
        for idx, value in enumerate(header_row)
        if value is not None and str(value).strip()
    }


def _require_columns(header_map: dict[str, int], required: set[str], sheet_name: str) -> None:
    missing = sorted(column for column in required if column not in header_map)
    if missing:
        raise ValueError(f"{sheet_name} is missing required columns: {', '.join(missing)}")


def _wallet_value(value: Any) -> str | None:
    text = str(value or "").strip().lower()
    return text if text.startswith("0x") else None


def _resolve_sheet_wallet(wallet_token: str | None, known_wallets: dict[str, dict[str, Any]]) -> str | None:
    token = _wallet_value(wallet_token)
    if not token:
        return None
    if token in known_wallets:
        return token
    matches = [wallet for wallet in known_wallets if wallet.startswith(token)]
    if len(matches) == 1:
        return matches[0]
    return None


def _normalize_copied_trade(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "trade_date": row["ts"].date(),
        "ts": row["ts"],
        "token_id": row["token_id"],
        "condition_id": row["condition_id"],
        "side": row["side"],
        "shares": float(row["copied_shares"] or 0.0),
        "usdc": float(row["copied_notional"] or 0.0),
        "fee_usdc": float(row.get("copied_fee") or 0.0),
        "price": float(row["copied_price"] or 0.0),
        "role": "sim",
        "question": row.get("question", ""),
    }


def _build_validation_payload(normalized_range: str, workbook_value: float | None, recomputed_value: float | None) -> dict[str, Any]:
    return {
        "workbook_value": workbook_value,
        "recomputed_value": recomputed_value,
        "delta": None if workbook_value is None or recomputed_value is None else round(recomputed_value - float(workbook_value), 2),
        "range_note": "" if normalized_range in {"1D", "7D", "30D"} else "Workbook provides validation only for 1D / 7D / 30D",
    }


def _token_scope_from_trades(
    trades: list[dict[str, Any]],
    opening_positions: dict[str, float],
    visible_counts: dict[str, int],
) -> list[dict[str, Any]]:
    ordered: dict[str, dict[str, Any]] = {}
    for trade in trades:
        ordered.setdefault(
            trade["token_id"],
            {
                "token_id": trade["token_id"],
                "condition_id": trade["condition_id"],
                "question": trade.get("question", ""),
                "first_trade_ts": trade["ts"],
                "last_trade_ts": trade["ts"],
                "opening_shares": opening_positions.get(trade["token_id"], 0.0),
                "visible_trade_count": visible_counts.get(trade["token_id"], 0),
            },
        )
        ordered[trade["token_id"]]["last_trade_ts"] = trade["ts"]
    return list(ordered.values())


def _latest_close_before_or_on(
    closes: list[dict[str, Any]],
    token_id: str,
    as_of_date: date,
) -> float | None:
    latest_price = None
    latest_date = None
    for close in closes:
        if close["token_id"] != token_id:
            continue
        trade_date = close["trade_date"]
        if trade_date > as_of_date:
            continue
        if latest_date is None or trade_date > latest_date:
            latest_date = trade_date
            latest_price = float(close["close_price"])
    return latest_price


def _augment_closes_for_opening_positions(
    closes: list[dict[str, Any]],
    token_scope: list[dict[str, Any]],
    window_start_date: date,
) -> list[dict[str, Any]]:
    baseline_date = window_start_date - timedelta(days=1)
    augmented = list(closes)
    for token in token_scope:
        if abs(float(token.get("opening_shares", 0.0) or 0.0)) <= 1e-9:
            continue
        already_has_baseline = any(
            close["token_id"] == token["token_id"] and close["trade_date"] == baseline_date
            for close in closes
        )
        if already_has_baseline:
            continue
        latest_price = _latest_close_before_or_on(closes, token["token_id"], baseline_date)
        if latest_price is None:
            continue
        augmented.append(
            {
                "token_id": token["token_id"],
                "trade_date": baseline_date,
                "close_price": latest_price,
            }
        )
    return augmented


def build_sim_payload(
    wallet_meta: dict[str, Any],
    drl_rows: list[dict[str, Any]],
    closes: list[dict[str, Any]],
    resolutions: dict[str, Any],
    range_key: Any = CURATION_ALL_RANGE,
) -> dict[str, Any]:
    normalized_range = normalize_curation_range_key(range_key)
    copied_rows = [row for row in drl_rows if row.get("status") == "COPIED"]
    if not copied_rows:
        return {
            "series": [],
            "summary": {
                "sim_status": wallet_meta.get("sim_status", "missing_drl"),
                "copied_trades": 0,
                "total_trades": 0,
                "final_pnl": 0.0,
                "roi_pct": 0.0,
            },
            "validation": {
                **_build_validation_payload(normalized_range, None, None),
            },
        }

    today_value = date.today()
    if normalized_range == CURATION_ALL_RANGE:
        window_start_date = min(row["ts"].date() for row in copied_rows)
    else:
        window_start_date = today_value - timedelta(days=CURATION_RANGE_DAYS[normalized_range])

    opening_positions: dict[str, float] = {}
    visible_trades: list[dict[str, Any]] = []
    visible_counts: dict[str, int] = {}
    normalized_trades = []

    for row in copied_rows:
        trade = _normalize_copied_trade(row)
        normalized_trades.append(trade)
        token_id = trade["token_id"]
        if trade["trade_date"] < window_start_date:
            signed_shares = trade["shares"] if trade["side"] == "BUY" else -trade["shares"]
            opening_positions[token_id] = opening_positions.get(token_id, 0.0) + signed_shares
            continue
        visible_trades.append(trade)
        visible_counts[token_id] = visible_counts.get(token_id, 0) + 1

    token_scope = _token_scope_from_trades(normalized_trades, opening_positions, visible_counts)
    scoped_closes = _augment_closes_for_opening_positions(closes, token_scope, window_start_date)
    chart = build_chart_payload(
        wallet_meta["address"],
        wallet_meta.get("filter_value", ""),
        CURATION_RANGE_DAYS.get(normalized_range, 0),
        token_scope,
        visible_trades,
        scoped_closes,
        resolutions,
        opening_positions=opening_positions,
        window_start_date=window_start_date,
    )
    if not chart:
        return {
            "series": [],
            "summary": {
                "sim_status": wallet_meta.get("sim_status", "missing_drl"),
                "copied_trades": len(copied_rows),
                "total_trades": len(copied_rows),
                "final_pnl": 0.0,
                "roi_pct": 0.0,
            },
            "validation": {
                **_build_validation_payload(normalized_range, None, 0.0),
            },
        }

    workbook_map = {
        "1D": wallet_meta.get("sim_1d"),
        "7D": wallet_meta.get("sim_7d"),
        "30D": wallet_meta.get("sim_30d"),
    }
    workbook_value = workbook_map.get(normalized_range)
    recomputed_value = chart["summary"]["final_pnl"]
    visible_volume = sum(trade["usdc"] for trade in visible_trades)
    chart["summary"]["copied_trades"] = len(copied_rows)
    chart["summary"]["sim_status"] = wallet_meta.get("sim_status", "ready")
    chart["summary"]["roi_pct"] = round((recomputed_value / visible_volume) * 100, 2) if visible_volume else 0.0
    chart["validation"] = _build_validation_payload(normalized_range, workbook_value, recomputed_value)
    return chart


def parse_sharpsim(file_bytes: bytes, filename: str = "") -> dict[str, Any]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    if "📊 Results" not in workbook.sheetnames:
        raise ValueError("Sharpsim workbook is missing the 📊 Results")

    results_ws = workbook["📊 Results"]
    results_rows = results_ws.iter_rows(values_only=True)
    results_header = _header_map(next(results_rows))
    _require_columns(results_header, _RESULTS_REQUIRED, "📊 Results")

    session_meta = {"filename": filename}
    if "📦 Portfolio" in workbook.sheetnames:
        portfolio_ws = workbook["📦 Portfolio"]
        for label, value, *_ in portfolio_ws.iter_rows(min_row=1, values_only=True):
            if str(label or "").strip() == "Capital (CLI)":
                session_meta["capital"] = float(value or 0.0)
                break

    wallets: dict[str, dict[str, Any]] = {}
    wallet_order: list[str] = []
    filter_summary: Counter[str] = Counter()

    for row in results_rows:
        wallet = _wallet_value(row[results_header["Wallet Address"]])
        if not wallet:
            continue
        filter_value = str(row[results_header["Detail"]] or "").strip()
        wallets[wallet] = {
            "address": wallet,
            "filter_level": "detail",
            "filter_value": filter_value,
            "category": str(row[results_header["Category"]] or "").strip(),
            "subcategory": str(row[results_header["Subcategory"]] or "").strip(),
            "total_trades": int(row[results_header["Trades"]] or 0),
            "volume": float(row[results_header["Volume"]] or 0.0),
            "sim_1d": float(row[results_header["💰 Sim 1d"]] or 0.0),
            "sim_7d": float(row[results_header["💰 Sim 7d"]] or 0.0),
            "sim_30d": float(row[results_header["💰 Sim 30d"]] or 0.0),
            "sim_roi_pct": float(row[results_header["📈 Sim ROI %"]] or 0.0),
            "max_dd_pct": float(row[results_header["📉 Max DD %"]] or 0.0),
            "copied": int(row[results_header["✅ Copied"]] or 0),
            "skipped": int(row[results_header["⏭️ Skipped"]] or 0),
            "sim_status": "missing_drl",
            "sim_error": "",
        }
        wallet_order.append(wallet)
        filter_summary[filter_value] += 1

    if not wallet_order:
        raise ValueError("Sharpsim workbook did not contain any valid wallet rows")

    drl: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith("_DRL"):
            continue
        drl_ws = workbook[sheet_name]
        rows = drl_ws.iter_rows(min_row=5, values_only=True)
        header_map = _header_map(next(rows))
        _require_columns(header_map, _DRL_REQUIRED, sheet_name)
        parts = sheet_name.split("_")
        wallet = _resolve_sheet_wallet(parts[1] if len(parts) > 1 else None, wallets)
        if not wallet:
            continue
        wallet_rows = []
        for row in rows:
            if not row or row[header_map["Timestamp (UTC)"]] is None:
                continue
            wallet_rows.append(
                {
                    "ts": row[header_map["Timestamp (UTC)"]],
                    "status": str(row[header_map["Status"]] or "").strip().upper(),
                    "side": str(row[header_map["Side"]] or "").strip().upper(),
                    "condition_id": str(row[header_map["Market ID"]] or "").strip(),
                    "question": str(row[header_map["Question"]] or "").strip(),
                    "token_id": str(row[header_map["Token ID"]] or "").strip(),
                    "copied_price": float(row[header_map["Copied Price"]] or 0.0),
                    "copied_shares": float(row[header_map["Copied Shares"]] or 0.0),
                    "copied_notional": float(row[header_map["Copied Notional $"]] or 0.0),
                    "copied_fee": float(row[header_map["Copied Fee $"]] or 0.0),
                }
            )
        drl[wallet] = wallet_rows
        if wallet in wallets:
            wallets[wallet]["sim_status"] = "ready"

    return {
        "session_meta": session_meta,
        "wallet_order": wallet_order,
        "wallets": wallets,
        "drl": drl,
        "filter_summary": dict(filter_summary),
        "parse_errors": [],
    }
