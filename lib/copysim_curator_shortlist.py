import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DETAIL_ORDER = [
    "League of Legends",
    "Counter-Strike",
    "Valorant",
    "Esports",
    "ALL_ESPORTS",
]

FLOOR_MINIMUMS = {
    "League of Legends": {"unique_events": 12, "unique_markets": 18, "unique_outcomes": 32, "copied": 100},
    "Counter-Strike": {"unique_events": 15, "unique_markets": 20, "unique_outcomes": 36, "copied": 100},
    "Valorant": {"unique_events": 8, "unique_markets": 10, "unique_outcomes": 18, "copied": 50},
    "Esports": {"unique_events": 8, "unique_markets": 10, "unique_outcomes": 20, "copied": 50},
    "ALL_ESPORTS": {"unique_events": 10, "unique_markets": 14, "unique_outcomes": 25, "copied": 68},
}

TIER_COLORS = {
    "Check First": "C6EFCE",
    "Check Next": "FFEB9C",
    "Deep Cut": "D9EAF7",
    "Monitor": "FCE4D6",
}


def _pct_rank(series: pd.Series, ascending: bool = True) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce").fillna(0.0)
    return numeric.rank(method="average", pct=True, ascending=ascending)


def _detail_thresholds(sub: pd.DataFrame, detail: str) -> dict[str, int]:
    q75 = sub[["unique_events", "unique_markets", "unique_outcomes", "copied"]].quantile(0.75)
    floors = FLOOR_MINIMUMS[detail]
    return {
        "unique_events": max(floors["unique_events"], int(q75["unique_events"])),
        "unique_markets": max(floors["unique_markets"], int(q75["unique_markets"])),
        "unique_outcomes": max(floors["unique_outcomes"], int(q75["unique_outcomes"])),
        "copied": max(floors["copied"], int(q75["copied"])),
    }


def _reason_columns(sub: pd.DataFrame) -> pd.DataFrame:
    q90_markets = sub["unique_markets"].quantile(0.90)
    q90_events = sub["unique_events"].quantile(0.90)
    q90_outcomes = sub["unique_outcomes"].quantile(0.90)
    q90_copied = sub["copied"].quantile(0.90)
    q90_pnl = sub["sim_pnl"].quantile(0.90)
    q25_skip = sub["skip_rate"].quantile(0.25)
    q75_skip = sub["skip_rate"].quantile(0.75)
    q75_dd = sub["max_dd_pct"].quantile(0.75)

    why_check: list[str] = []
    watch_items: list[str] = []
    for row in sub.itertuples(index=False):
        reasons: list[str] = []
        risks: list[str] = []
        if row.unique_markets >= q90_markets:
            reasons.append("elite market breadth")
        elif row.unique_markets >= sub["unique_markets"].quantile(0.75):
            reasons.append("strong market breadth")
        if row.unique_events >= q90_events:
            reasons.append("broad event coverage")
        if row.unique_outcomes >= q90_outcomes:
            reasons.append("deep outcome coverage")
        if row.copied >= q90_copied:
            reasons.append("high copied volume")
        if row.sim_pnl >= q90_pnl:
            reasons.append("strong sim pnl")
        if row.skip_rate <= q25_skip:
            reasons.append("low skipped-trade rate")
        if row.sim_rank > 300:
            reasons.append("deep cut beyond top 300")
        if not reasons:
            reasons.append("broad enough to validate manually")

        if row.skip_rate >= q75_skip:
            risks.append("high skipped-trade rate")
        if row.max_dd_pct >= q75_dd:
            risks.append("high max drawdown")
        if row.cap_skips_30d > 0 or row.market_cap_skips_30d > 0:
            risks.append("cap-constrained in sim")
        if row.include == 1:
            risks.append("already included")
        if not risks:
            risks.append("no major workbook warning")

        why_check.append("; ".join(reasons[:3]))
        watch_items.append("; ".join(risks[:3]))

    sub = sub.copy()
    sub["why_check"] = why_check
    sub["watch_items"] = watch_items
    return sub


def _build_detail_shortlist(df: pd.DataFrame, detail: str) -> tuple[pd.DataFrame, dict[str, int]]:
    sub = df[df["Detail"] == detail].copy()
    thresholds = _detail_thresholds(sub, detail)

    sub["skip_rate"] = sub["skipped"] / (sub["copied"] + sub["skipped"]).replace(0, 1)
    sub["cap_skip_rate"] = sub["cap_skips_30d"] / sub["sim_trades"].replace(0, 1)
    sub["market_cap_skip_rate"] = sub["market_cap_skips_30d"] / sub["sim_trades"].replace(0, 1)

    breadth_score = (
        _pct_rank(sub["unique_markets"]) * 0.35
        + _pct_rank(sub["unique_events"]) * 0.25
        + _pct_rank(sub["unique_outcomes"]) * 0.15
        + _pct_rank(sub["copied"]) * 0.15
        + _pct_rank(sub["sim_trades"]) * 0.10
    )
    pnl_score = _pct_rank(sub["sim_pnl"]) * 0.75 + _pct_rank(sub["sim_30d"]) * 0.25
    friction_score = (
        _pct_rank(sub["skip_rate"], ascending=False) * 0.50
        + _pct_rank(sub["market_cap_skip_rate"], ascending=False) * 0.25
        + _pct_rank(sub["cap_skip_rate"], ascending=False) * 0.15
        + _pct_rank(sub["max_dd_pct"], ascending=False) * 0.10
    )

    sub["breadth_score"] = breadth_score * 100
    sub["profit_score"] = pnl_score * 100
    sub["friction_score"] = friction_score * 100
    sub["review_score"] = (breadth_score * 0.60 + pnl_score * 0.25 + friction_score * 0.15) * 100

    eligible = sub[
        (sub["sim_pnl"] > 0)
        & (sub["include"] == 0)
        & (sub["unique_events"] >= thresholds["unique_events"])
        & (sub["unique_markets"] >= thresholds["unique_markets"])
        & (sub["unique_outcomes"] >= thresholds["unique_outcomes"])
        & (sub["copied"] >= thresholds["copied"])
    ].copy()

    check_first_cut = eligible["review_score"].quantile(0.80) if not eligible.empty else 0
    check_next_cut = eligible["review_score"].quantile(0.60) if not eligible.empty else 0
    deep_cut_cut = eligible["review_score"].quantile(0.70) if not eligible.empty else 0

    def assign_tier(row: pd.Series) -> str:
        if row["sim_rank"] > 300 and row["review_score"] >= deep_cut_cut:
            return "Deep Cut"
        if row["review_score"] >= check_first_cut:
            return "Check First"
        if row["review_score"] >= check_next_cut:
            return "Check Next"
        return "Monitor"

    eligible["review_tier"] = eligible.apply(assign_tier, axis=1)
    eligible["deep_cut"] = eligible["sim_rank"] > 300
    eligible = _reason_columns(eligible)
    eligible = eligible[eligible["review_tier"].isin(["Check First", "Check Next", "Deep Cut"])]
    eligible = eligible.sort_values(
        ["review_tier", "review_score", "breadth_score", "sim_pnl"],
        ascending=[True, False, False, False],
    ).copy()
    eligible.insert(0, "curator_rank", range(1, len(eligible) + 1))
    return eligible, thresholds


def _load_results(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="📊 Results", dtype={"Wallet Address": str, "Unique E/M/O": str})
    ueo = df["Unique E/M/O"].fillna("0/0/0").astype(str).str.split("/", expand=True)
    df["unique_events"] = pd.to_numeric(ueo[0], errors="coerce").fillna(0).astype(int)
    df["unique_markets"] = pd.to_numeric(ueo[1], errors="coerce").fillna(0).astype(int)
    df["unique_outcomes"] = pd.to_numeric(ueo[2], errors="coerce").fillna(0).astype(int)

    rename_map = {
        "Wallet Address": "wallet_address",
        "Include": "include",
        "Category": "category",
        "Subcategory": "subcategory",
        "Detail": "detail",
        "Trades": "trades",
        "Sim Trades": "sim_trades",
        "Volume": "volume",
        "Raw PnL": "raw_pnl",
        "💰 Sim 30d": "sim_30d",
        "⛔ Cap Skips 30d": "cap_skips_30d",
        "⛔ Mkt Cap Skips 30d": "market_cap_skips_30d",
        "💰 Sim PnL": "sim_pnl",
        "📈 Sim ROI %": "sim_roi_pct",
        "📉 Max DD %": "max_dd_pct",
        "✅ Copied": "copied",
        "⏭️ Skipped": "skipped",
        "⚠️ Flag": "flag",
        "Execution Mode": "execution_mode",
        "Sim Rank": "sim_rank",
        "Official ROI %": "official_roi_pct",
        "Official PnL": "official_pnl",
        "0.995 Cap ROI %": "cap_0995_roi_pct",
        "0.995 Cap PnL": "cap_0995_pnl",
    }
    return df.rename(columns=rename_map)


def _export_shortlist(
    output_path: Path,
    source_path: Path,
    detail_results: dict[str, pd.DataFrame],
    thresholds_by_detail: dict[str, dict[str, int]],
    source_rows: int,
) -> None:
    summary_rows = []
    for detail in DETAIL_ORDER:
        result = detail_results.get(detail)
        if result is None:
            continue
        thresholds = thresholds_by_detail[detail]
        summary_rows.append(
            {
                "Detail": detail,
                "Shortlisted": len(result),
                "Check First": int((result["review_tier"] == "Check First").sum()),
                "Check Next": int((result["review_tier"] == "Check Next").sum()),
                "Deep Cut": int((result["review_tier"] == "Deep Cut").sum()),
                "Threshold Events": thresholds["unique_events"],
                "Threshold Markets": thresholds["unique_markets"],
                "Threshold Outcomes": thresholds["unique_outcomes"],
                "Threshold Copied": thresholds["copied"],
            }
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=6)
        workbook = writer.book
        sheet = writer.sheets["Summary"]
        sheet["A1"] = "Wallet Curator Shortlist"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A2"] = f"Source workbook: {source_path.name}"
        sheet["A3"] = f"Source rows scanned: {source_rows:,}"
        sheet["A4"] = "This shortlist favors breadth and repeatability over headline ROI."
        sheet["A5"] = "Eligibility requires include=0, positive sim pnl, and minimum unique events, markets, outcomes, and copied-trade depth by detail."

        all_candidates = []
        for detail in DETAIL_ORDER:
            result = detail_results.get(detail)
            if result is None or result.empty:
                continue
            export = result[
                [
                    "curator_rank",
                    "review_tier",
                    "review_score",
                    "breadth_score",
                    "profit_score",
                    "friction_score",
                    "wallet_address",
                    "include",
                    "sim_rank",
                    "deep_cut",
                    "sim_pnl",
                    "sim_roi_pct",
                    "sim_30d",
                    "max_dd_pct",
                    "copied",
                    "skipped",
                    "skip_rate",
                    "unique_events",
                    "unique_markets",
                    "unique_outcomes",
                    "trades",
                    "sim_trades",
                    "volume",
                    "cap_skips_30d",
                    "market_cap_skips_30d",
                    "execution_mode",
                    "flag",
                    "why_check",
                    "watch_items",
                ]
            ].copy()
            export.insert(7, "detail", detail)
            export.to_excel(writer, sheet_name=detail[:31], index=False)
            all_candidates.append(export)

        if all_candidates:
            combined = pd.concat(all_candidates, ignore_index=True).sort_values(
                ["detail", "review_tier", "review_score"],
                ascending=[True, True, False],
            )
            combined.to_excel(writer, sheet_name="All Shortlist", index=False)
            deep_cuts = combined[combined["deep_cut"]].sort_values(
                ["detail", "review_score", "breadth_score"],
                ascending=[True, False, False],
            )
            if not deep_cuts.empty:
                deep_cuts.to_excel(writer, sheet_name="Deep Cuts", index=False)


def _style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        header_row = 1 if ws.title != "Summary" else 7
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1F2937")
        if "review_tier" in [c.value for c in ws[header_row]]:
            header = [c.value for c in ws[header_row]]
            tier_col = header.index("review_tier") + 1
            for row in range(header_row + 1, ws.max_row + 1):
                value = ws.cell(row=row, column=tier_col).value
                fill = TIER_COLORS.get(value)
                if fill:
                    ws.cell(row=row, column=tier_col).fill = PatternFill("solid", fgColor=fill)
        for col_cells in ws.columns:
            values = [str(cell.value) for cell in col_cells if cell.value is not None]
            if not values:
                continue
            width = min(max(len(v) for v in values) + 2, 36)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width
    wb.save(path)


def build_shortlist(input_path: Path, output_path: Path) -> Path:
    df = _load_results(input_path)
    source_rows = len(df)
    detail_results: dict[str, pd.DataFrame] = {}
    thresholds_by_detail: dict[str, dict[str, int]] = {}

    for detail in DETAIL_ORDER:
        if detail not in set(df["detail"]):
            continue
        result, thresholds = _build_detail_shortlist(
            df.rename(columns={"detail": "Detail"}), detail
        )
        detail_results[detail] = result
        thresholds_by_detail[detail] = thresholds

    _export_shortlist(output_path, input_path, detail_results, thresholds_by_detail, source_rows)
    _style_workbook(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a breadth-first Wallet Curator shortlist from a Copysim workbook.")
    parser.add_argument("input_path", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    output_path = args.output or Path("reports") / f"{args.input_path.stem} Wallet Curator Shortlist.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    path = build_shortlist(args.input_path, output_path)
    print(path)


if __name__ == "__main__":
    main()
